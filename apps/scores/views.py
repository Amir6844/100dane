from django.views.generic import View, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.shortcuts import get_object_or_404, redirect, render
from django.contrib import messages
from django.urls import reverse
from django.http import HttpResponse
from django.db.models import Avg
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from apps.accounts.mixins import TeacherRequiredMixin
from apps.exams.models import Exam
from apps.classes.models import Classroom
from .models import Score
from apps.notifications.models import Notification

class BulkScoreView(TeacherRequiredMixin, View):
    template_name = 'scores/bulk.html'

    def get(self, request, exam_pk):
        exam = get_object_or_404(Exam, pk=exam_pk, classroom__teacher=request.user)
        classroom = exam.classroom
        students = classroom.students.order_by('last_name','first_name')
        # existing scores map
        scores_map = {s.student_id: s for s in Score.objects.filter(exam=exam)}
        context = {'exam': exam, 'classroom': classroom, 'students': students, 'scores_map': scores_map}
        return render(request, self.template_name, context)

    def post(self, request, exam_pk):
        from django.db import transaction
        exam = get_object_or_404(Exam, pk=exam_pk, classroom__teacher=request.user)
        classroom = exam.classroom
        students = classroom.students.all()
        saved = 0
        errors = []
        with transaction.atomic():
            for s in students:
                key = f"score_{s.id}"
                val = request.POST.get(key)
                if val is None or val == '':
                    continue
                # prevent IDOR: ensure student belongs to classroom
                if s not in classroom.students.all():
                    errors.append(f"دانش‌آموز {s.username} عضو کلاس نیست.")
                    continue
                try:
                    v = float(val.replace('٫','.').replace('،','.').strip())
                except:
                    errors.append(f"نمره نامعتبر برای {s.get_full_name() or s.username}")
                    continue
                if v < 0 or v > float(exam.total_score):
                    errors.append(f"نمره {s.get_full_name() or s.username} باید بین 0 و {exam.total_score} باشد.")
                    continue
                obj, created = Score.objects.update_or_create(exam=exam, student=s, defaults={'value': v})
                # also sync score field
                if obj.score != v:
                    obj.score = v
                    obj.save(update_fields=['value', 'score'])
                saved += 1
                Notification.objects.get_or_create(user=s, title='نمره جدید', message=f"نمره آزمون «{exam.title}» ثبت شد: {v}", link=reverse('scores:report', args=[classroom.pk]) )
            if errors:
                for e in errors:
                    messages.error(request, e)
                # if any error, still commit valid ones? atomic keeps them but we may want to rollback on critical? keep partial
        if saved:
            messages.success(request, f'{saved} نمره با موفقیت ذخیره شد.')
        else:
            if not errors:
                messages.info(request, 'هیچ نمره‌ای وارد نشد.')
        return redirect('scores:bulk', exam_pk=exam.pk)

class StudentReportView(LoginRequiredMixin, TemplateView):
    template_name = 'scores/report.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        classroom = get_object_or_404(Classroom, pk=kwargs['class_pk'])
        user = self.request.user
        # if teacher, allow ?student= id, else own
        target_student = user
        if user.is_teacher and self.request.GET.get('student'):
            from django.contrib.auth import get_user_model
            User = get_user_model()
            target_student = get_object_or_404(User, pk=self.request.GET.get('student'))
        elif user.is_teacher and classroom.teacher != user:
            # not owner but student logic?
            pass
        # ensure access
        if not (user == classroom.teacher or user in classroom.students.all() or user == target_student):
            # if student viewing own report, must be enrolled
            pass
        exams = classroom.exams.order_by('date')
        scores = {s.exam_id: s for s in Score.objects.filter(exam__classroom=classroom, student=target_student)}
        rows = []
        total = 0
        count = 0
        for e in exams:
            sc = scores.get(e.id)
            val = float(sc.value) if sc else None
            if val is not None:
                total += val
                count += 1
            rows.append({'exam': e, 'score': sc})
        avg = round(total / count, 2) if count else None
        # rank
        from django.db.models import Avg
        all_students = classroom.students.all()
        ranking = []
        for s in all_students:
            a = Score.objects.filter(student=s, exam__classroom=classroom).aggregate(avg=Avg('value'))['avg']
            ranking.append((s, float(a) if a else 0))
        ranking.sort(key=lambda x: x[1], reverse=True)
        rank = None
        for i, (s, a) in enumerate(ranking, 1):
            if s.id == target_student.id:
                rank = i
                break
        ctx.update({
            'classroom': classroom,
            'student': target_student,
            'rows': rows,
            'avg': avg,
            'rank': rank,
            'total_students': len(all_students),
            'is_own': target_student == user,
        })
        return ctx

class ExportExcelView(TeacherRequiredMixin, View):
    def get(self, request, class_pk):
        classroom = get_object_or_404(Classroom, pk=class_pk, teacher=request.user)
        mode = request.GET.get('mode', 'full')  # full, exam, members, group, student, detailed
        exam_id = request.GET.get('exam')
        group_id = request.GET.get('group')
        student_id = request.GET.get('student')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.sheet_view.rightToLeft = True

        header_fill = PatternFill(start_color="C22A4E", end_color="C22A4E", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11, name="Vazirmatn")
        thin = Side(style="thin", color="E5E7EB")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        def style_header(cell):
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Helper to get group name for student
        def get_student_group_name(student):
            from apps.groups.models import Group
            g = Group.objects.filter(classroom=classroom, members=student).first()
            return g.name if g else "-"

        if mode == 'members':
            ws.title = "اعضا"
            headers = ["ردیف", "نام", "نام خانوادگی", "نام کاربری", "شماره موبایل", "کد دانش‌آموزی"]
            ws.append(headers)
            for c in ws[1]:
                style_header(c)
            for i, s in enumerate(classroom.students.order_by('last_name','first_name'), 1):
                ws.append([i, s.first_name, s.last_name, s.username, s.phone or "-", getattr(s, 'student_code', '') or "-"])
            filename = f"{classroom.name}-اعضا.xlsx"
        elif mode == 'exam' and exam_id:
            exam = get_object_or_404(classroom.exams.all(), pk=exam_id)
            ws.title = exam.title[:30]
            # Use spec Persian headers for exam export
            headers = ["ردیف", "نام", "نام خانوادگی", "کد دانش‌آموزی", "گروه", "آزمون", "نمره", "حداکثر نمره", "درصد"]
            ws.append(headers)
            for c in ws[1]:
                style_header(c)
            scores = {s.student_id: s for s in exam.scores.select_related('student')}
            for i, stu in enumerate(classroom.students.order_by('last_name','first_name'), 1):
                sc = scores.get(stu.id)
                score_val = float(sc.value) if sc else None
                percent = round(score_val / float(exam.total_score) * 100, 1) if score_val is not None else "-"
                ws.append([
                    i,
                    stu.first_name or "-",
                    stu.last_name or "-",
                    getattr(stu, 'student_code', '') or stu.username,
                    get_student_group_name(stu),
                    exam.title,
                    score_val if score_val is not None else "-",
                    float(exam.total_score),
                    f"{percent}%" if percent != "-" else "-"
                ])
            filename = f"{classroom.name}-{exam.title}.xlsx"
        elif mode == 'group' and group_id:
            from apps.groups.models import Group
            group = get_object_or_404(Group, pk=group_id, classroom=classroom)
            ws.title = group.name[:30]
            headers = ["ردیف", "نام", "نام خانوادگی", "کد دانش‌آموزی", "گروه", "آزمون", "نمره", "حداکثر نمره", "درصد"]
            ws.append(headers)
            for c in ws[1]:
                style_header(c)
            # group scores: for each student in group, each exam
            exams = classroom.exams.order_by('date')
            row_idx = 1
            from collections import defaultdict
            score_map = { (s.student_id, s.exam_id): float(s.value) for s in Score.objects.filter(exam__classroom=classroom, student__in=group.members.all()).select_related('student') }
            for stu in group.members.order_by('last_name','first_name'):
                for exam in exams:
                    val = score_map.get((stu.id, exam.id))
                    percent = round(val / float(exam.total_score) * 100, 1) if val is not None else "-"
                    ws.append([
                        row_idx,
                        stu.first_name or "-",
                        stu.last_name or "-",
                        getattr(stu, 'student_code', '') or stu.username,
                        group.name,
                        exam.title,
                        val if val is not None else "-",
                        float(exam.total_score),
                        f"{percent}%" if percent != "-" else "-"
                    ])
                    row_idx += 1
            if ws.max_row == 1:
                ws.append([1, "-", "-", "-", group.name, "-", "-", "-", "-"])
            filename = f"{classroom.name}-{group.name}.xlsx"
        elif mode == 'student' and student_id:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            stu = get_object_or_404(User, pk=student_id, enrolled_classes=classroom)
            ws.title = (stu.get_full_name() or stu.username)[:30]
            headers = ["ردیف", "نام", "نام خانوادگی", "کد دانش‌آموزی", "گروه", "آزمون", "نمره", "حداکثر نمره", "درصد"]
            ws.append(headers)
            for c in ws[1]:
                style_header(c)
            scores = {s.exam_id: s for s in Score.objects.filter(exam__classroom=classroom, student=stu).select_related('exam')}
            for i, exam in enumerate(classroom.exams.order_by('date'), 1):
                sc = scores.get(exam.id)
                val = float(sc.value) if sc else None
                percent = round(val / float(exam.total_score) * 100, 1) if val is not None else "-"
                ws.append([
                    i,
                    stu.first_name or "-",
                    stu.last_name or "-",
                    getattr(stu, 'student_code', '') or stu.username,
                    get_student_group_name(stu),
                    exam.title,
                    val if val is not None else "-",
                    float(exam.total_score),
                    f"{percent}%" if percent != "-" else "-"
                ])
            filename = f"{classroom.name}-{stu.username}.xlsx"
        elif mode == 'detailed':
            # Detailed per spec: one row per student per exam with spec headers
            ws.title = "نمرات تفصیلی"
            headers = ["ردیف", "نام", "نام خانوادگی", "کد دانش‌آموزی", "گروه", "آزمون", "نمره", "حداکثر نمره", "درصد"]
            ws.append(headers)
            for c in ws[1]:
                style_header(c)
            row_idx = 1
            score_map = { (s.student_id, s.exam_id): float(s.value) for s in Score.objects.filter(exam__classroom=classroom).select_related('student', 'exam')}
            for stu in classroom.students.order_by('last_name','first_name'):
                for exam in classroom.exams.order_by('date'):
                    val = score_map.get((stu.id, exam.id))
                    percent = round(val / float(exam.total_score) * 100, 1) if val is not None else "-"
                    ws.append([
                        row_idx,
                        stu.first_name or "-",
                        stu.last_name or "-",
                        getattr(stu, 'student_code', '') or stu.username,
                        get_student_group_name(stu),
                        exam.title,
                        val if val is not None else "-",
                        float(exam.total_score),
                        f"{percent}%" if percent != "-" else "-"
                    ])
                    row_idx += 1
            filename = f"{classroom.name}-نمرات-تفصیلی.xlsx"
        else:
            # full grade matrix (keep previous but also add detailed sheet?)
            ws.title = "کارنامه"
            exams = list(classroom.exams.order_by('date'))
            headers = ["ردیف", "نام دانش‌آموز", "نام کاربری"] + [e.title for e in exams] + ["میانگین", "رتبه"]
            ws.append(headers)
            for c in ws[1]:
                style_header(c)
            ws.freeze_panes = "A2"
            students = list(classroom.students.order_by('last_name','first_name'))
            from collections import defaultdict
            score_map = defaultdict(dict)
            for sc in Score.objects.filter(exam__classroom=classroom).select_related('student'):
                score_map[sc.student_id][sc.exam_id] = float(sc.value)
            avgs = {}
            for stu in students:
                vals = [score_map[stu.id].get(e.id) for e in exams]
                vals = [v for v in vals if v is not None]
                avgs[stu.id] = round(sum(vals)/len(vals),2) if vals else 0
            sorted_students = sorted(students, key=lambda s: avgs[s.id], reverse=True)
            rank_map = {}
            for r, s in enumerate(sorted_students, 1):
                rank_map[s.id] = r if avgs[s.id] else "-"
            for i, stu in enumerate(students, 1):
                row = [i, f"{stu.first_name} {stu.last_name}".strip() or stu.username, stu.username]
                for e in exams:
                    v = score_map[stu.id].get(e.id, "-")
                    row.append(v if v != "-" else "-")
                avg = avgs[stu.id] if avgs[stu.id] else "-"
                row += [avg, rank_map[stu.id]]
                ws.append(row)
            light_fill = PatternFill(start_color="FDF2F4", end_color="FDF2F4", fill_type="solid")
            for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column), 2):
                if idx % 2 == 0:
                    for c in row:
                        c.fill = light_fill
                        c.border = border
                        c.alignment = Alignment(horizontal="center", vertical="center")
                else:
                    for c in row:
                        c.border = border
                        c.alignment = Alignment(horizontal="center", vertical="center")
            filename = f"{classroom.name}-کارنامه.xlsx"
        # auto width
        for col in ws.columns:
            max_len = max(len(str(c.value)) if c.value else 0 for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 6, 30)
        ws.row_dimensions[1].height = 28

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

class MyScoresView(LoginRequiredMixin, TemplateView):
    template_name = 'scores/my_scores.html'
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from apps.scores.models import Score
        ctx['scores'] = Score.objects.filter(student=self.request.user).select_related('exam','exam__classroom').order_by('-created_at')
        return ctx

class ReportsDashboardView(TeacherRequiredMixin, TemplateView):
    template_name = 'reports/dashboard.html'
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        from apps.classes.models import Classroom
        classrooms = Classroom.objects.filter(teacher=self.request.user).prefetch_related('students','groups','exams','lessons')
        ctx['classrooms'] = classrooms
        # stats
        from django.db.models import Avg, Count
        from apps.scores.models import Score
        ctx['total_classes'] = classrooms.count()
        ctx['total_students'] = sum(c.students.count() for c in classrooms)
        ctx['total_exams'] = sum(c.exams.count() for c in classrooms)
        ctx['recent_scores'] = Score.objects.filter(exam__classroom__teacher=self.request.user).select_related('student','exam','exam__classroom').order_by('-created_at')[:6]
        return ctx

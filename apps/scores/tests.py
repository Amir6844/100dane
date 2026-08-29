from django.test import TestCase, Client
from django.urls import reverse
from django.contrib.auth import get_user_model
from apps.classes.models import Classroom
from apps.exams.models import Exam
from apps.scores.models import Score
from apps.groups.models import Group
import io
import openpyxl

User = get_user_model()

class ScoreTests(TestCase):
    def setUp(self):
        self.teacher = User.objects.create_user(username='t', password='p', role=User.Role.TEACHER)
        self.student = User.objects.create_user(username='s', password='p', role=User.Role.STUDENT, first_name='علی', last_name='رضایی', student_code='001')
        self.student2 = User.objects.create_user(username='s2', password='p', role=User.Role.STUDENT)
        self.outsider = User.objects.create_user(username='out', password='p', role=User.Role.STUDENT)
        self.classroom = Classroom.objects.create(name='کلاس', teacher=self.teacher)
        self.classroom.students.add(self.student, self.student2)
        self.exam = Exam.objects.create(classroom=self.classroom, title='آزمون', total_score=20)
        self.group = Group.objects.create(classroom=self.classroom, name='گروه')
        self.group.members.add(self.student)

    def test_valid_score(self):
        s = Score(exam=self.exam, student=self.student, value=15)
        s.full_clean()
        s.save()
        self.assertEqual(Score.objects.count(), 1)

    def test_negative_score(self):
        from django.core.exceptions import ValidationError
        s = Score(exam=self.exam, student=self.student, value=-1)
        with self.assertRaises(ValidationError):
            s.full_clean()

    def test_exceed_max(self):
        from django.core.exceptions import ValidationError
        s = Score(exam=self.exam, student=self.student, value=25)
        with self.assertRaises(ValidationError):
            s.full_clean()

    def test_bulk_entry(self):
        self.client.login(username='t', password='p')
        resp = self.client.post(reverse('scores:bulk', args=[self.exam.pk]), {f'score_{self.student.id}': '18', f'score_{self.student2.id}': '19'})
        self.assertIn(resp.status_code, [302, 200])
        self.assertTrue(Score.objects.filter(student=self.student, value=18).exists())
        self.assertTrue(Score.objects.filter(student=self.student2, value=19).exists())

    def test_bulk_validation(self):
        self.client.login(username='t', password='p')
        self.client.post(reverse('scores:bulk', args=[self.exam.pk]), {f'score_{self.student.id}': '30'})  # exceeds max
        self.assertFalse(Score.objects.filter(student=self.student).exists())

    def test_unauthorized_bulk(self):
        t2 = User.objects.create_user(username='t2', password='p', role=User.Role.TEACHER)
        self.client.login(username='t2', password='p')
        resp = self.client.get(reverse('scores:bulk', args=[self.exam.pk]))
        self.assertEqual(resp.status_code, 404)

    def test_student_outside_class_not_scored(self):
        # outsider tries to get score via bulk - should be ignored because not in classroom.students.all loop
        self.client.login(username='t', password='p')
        # bulk only loops over classroom.students, so outsider won't be processed; manually try to create Score for outsider via bulk param shouldn't affect
        self.client.post(reverse('scores:bulk', args=[self.exam.pk]), {f'score_{self.outsider.id}': '15'})
        self.assertFalse(Score.objects.filter(student=self.outsider).exists())

    def test_excel_export_full(self):
        Score.objects.create(exam=self.exam, student=self.student, value=16)
        Score.objects.create(exam=self.exam, student=self.student2, value=12)
        self.client.login(username='t', password='p')
        resp = self.client.get(reverse('scores:export', args=[self.classroom.pk]))
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp['Content-Type'], 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        wb = openpyxl.load_workbook(filename=io.BytesIO(resp.content))
        ws = wb.active
        self.assertIsNotNone(ws)

    def test_excel_persian_headers(self):
        Score.objects.create(exam=self.exam, student=self.student, value=15)
        self.client.login(username='t', password='p')
        resp = self.client.get(reverse('scores:export', args=[self.classroom.pk]) + '?mode=exam&exam=' + str(self.exam.pk))
        wb = openpyxl.load_workbook(filename=io.BytesIO(resp.content))
        ws = wb.active
        headers = [c.value for c in ws[1]]
        # must contain spec headers
        self.assertIn('ردیف', headers)
        self.assertIn('نام', headers)
        self.assertIn('نام خانوادگی', headers)
        self.assertIn('کد دانش‌آموزی', headers)
        self.assertIn('گروه', headers)
        self.assertIn('آزمون', headers)
        self.assertIn('نمره', headers)
        self.assertIn('حداکثر نمره', headers)
        self.assertIn('درصد', headers)

    def test_excel_group_export(self):
        self.client.login(username='t', password='p')
        resp = self.client.get(reverse('scores:export', args=[self.classroom.pk]) + f'?mode=group&group={self.group.pk}')
        self.assertEqual(resp.status_code, 200)

    def test_excel_student_export(self):
        self.client.login(username='t', password='p')
        resp = self.client.get(reverse('scores:export', args=[self.classroom.pk]) + f'?mode=student&student={self.student.pk}')
        self.assertEqual(resp.status_code, 200)

    def test_statistics_services(self):
        from apps.scores.services import class_stats, student_stats, exam_stats, group_stats
        Score.objects.create(exam=self.exam, student=self.student, value=10)
        Score.objects.create(exam=self.exam, student=self.student2, value=20)
        cs = class_stats(self.classroom)
        self.assertEqual(cs['student_count'], 2)
        self.assertEqual(cs['exam_count'], 1)
        self.assertAlmostEqual(cs['average'], 15.0)
        ss = student_stats(self.student, self.classroom)
        self.assertEqual(ss['exam_count'], 1)
        self.assertEqual(ss['average'], 10)
        es = exam_stats(self.exam)
        self.assertEqual(es['participant_count'], 2)
        self.assertEqual(es['average'], 15)
        gs = group_stats(self.group)
        self.assertEqual(gs['member_count'], 1)

class StudentRosterTests(TestCase):
    def setUp(self):
        self.teacher = User.objects.create_user(username='t', password='p', role=User.Role.TEACHER)
        self.classroom = Classroom.objects.create(name='ک', teacher=self.teacher)

    def test_roster_create(self):
        self.client.login(username='t', password='p')
        resp = self.client.post(reverse('students:roster_create', args=[self.classroom.pk]), {'first_name': 'سارا', 'last_name': 'محمدی', 'student_code': '1001', 'phone': '09123456789'})
        self.assertIn(resp.status_code, [302, 200])
        from apps.students.models import Student
        self.assertTrue(Student.objects.filter(student_code='1001').exists())

    def test_duplicate_student_code(self):
        from apps.students.models import Student
        Student.objects.create(classroom=self.classroom, first_name='a', last_name='b', student_code='1001')
        self.client.login(username='t', password='p')
        resp = self.client.post(reverse('students:roster_create', args=[self.classroom.pk]), {'first_name': 'c', 'last_name': 'd', 'student_code': '1001'})
        self.assertEqual(resp.status_code, 200)
        self.assertContains(resp, 'تکراری')

    def test_enrolled_detail_stats(self):
        student = User.objects.create_user(username='s', password='p', role=User.Role.STUDENT, first_name='حسین')
        self.classroom.students.add(student)
        exam = Exam.objects.create(classroom=self.classroom, title='آزمون', total_score=20)
        Score.objects.create(exam=exam, student=student, value=18)
        self.client.login(username='t', password='p')
        resp = self.client.get(reverse('students:detail', args=[self.classroom.pk, student.pk]))
        self.assertEqual(resp.status_code, 200)
        self.assertIn('avg', resp.context)
